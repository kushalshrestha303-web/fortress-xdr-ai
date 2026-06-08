import csv
import io
import json
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook

from app.models.log_analyzer import ParsedLogEvent

IP_RE = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
TIME_RE = re.compile(r"(\d{4}-\d{2}-\d{2}[T\s]\d{2}:\d{2}:\d{2}(?:Z)?)")


def sanitize_log_text(text: str) -> str:
    return text.replace("\x00", "").replace("<", "&lt;").replace(">", "&gt;")


def parse_log_bytes(filename: str, content: bytes) -> list[ParsedLogEvent]:
    if filename.lower().endswith(".pcapng"):
        return [
            ParsedLogEvent(
                timestamp=datetime.now(timezone.utc).isoformat(),
                protocol="PCAPNG",
                severity="medium",
                info="Binary pcapng upload stored safely. Packet decoding requires a tshark/scapy worker integration.",
                raw="pcapng binary not executed",
            )
        ]

    text = sanitize_log_text(content.decode("utf-8", errors="replace"))
    if filename.lower().endswith(".json"):
        return _parse_json(text)
    if filename.lower().endswith(".csv"):
        return _parse_csv(text)
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_lines(text)


def _parse_json(text: str) -> list[ParsedLogEvent]:
    loaded = json.loads(text)
    records = loaded if isinstance(loaded, list) else loaded.get("events", [loaded])
    return [_event_from_mapping(record) for record in records if isinstance(record, dict)]


def _parse_csv(text: str) -> list[ParsedLogEvent]:
    reader = csv.DictReader(io.StringIO(text))
    return [_event_from_mapping(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[ParsedLogEvent]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []

    headers = [_normalize_header(value, index) for index, value in enumerate(header_row)]
    events: list[ParsedLogEvent] = []
    for row in rows:
        record = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and value not in (None, "")
        }
        if not record:
            continue
        events.append(_event_from_mapping(record))
    return events


def _parse_lines(text: str) -> list[ParsedLogEvent]:
    events: list[ParsedLogEvent] = []
    for line in text.splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue
        ips = IP_RE.findall(cleaned)
        protocol = _protocol_from_text(cleaned)
        events.append(
            ParsedLogEvent(
                timestamp=_timestamp_from_text(cleaned),
                source_ip=ips[0] if ips else None,
                destination_ip=ips[1] if len(ips) > 1 else None,
                protocol=protocol,
                info=_info_from_text(cleaned),
                severity=_severity_from_text(cleaned),
                raw=cleaned[:2000],
            )
        )
    return events


def _event_from_mapping(record: dict[str, Any]) -> ParsedLogEvent:
    raw = json.dumps(record, ensure_ascii=True)[:2000]
    source = _first(record, "source_ip", "src_ip", "src", "source", "Source")
    destination = _first(record, "destination_ip", "dst_ip", "dst", "destination", "Destination")
    info = str(_first(record, "info", "message", "Info", "event", "url") or raw)
    return ParsedLogEvent(
        timestamp=str(_first(record, "timestamp", "time", "Time", "@timestamp") or datetime.now(timezone.utc).isoformat()),
        source_ip=str(source) if source else None,
        destination_ip=str(destination) if destination else None,
        protocol=str(_first(record, "protocol", "proto", "Protocol") or _protocol_from_text(info)).upper(),
        info=sanitize_log_text(info[:2000]),
        severity=_severity_from_text(info),
        raw=sanitize_log_text(raw),
    )


def _first(record: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in record and record[key] not in (None, ""):
            return record[key]
    return None


def _normalize_header(value: Any, index: int) -> str:
    if value is None:
        return f"column_{index + 1}"
    cleaned = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")
    aliases = {
        "source": "source_ip",
        "source_ip_address": "source_ip",
        "src": "source_ip",
        "src_ip": "source_ip",
        "destination": "destination_ip",
        "destination_ip_address": "destination_ip",
        "dst": "destination_ip",
        "dst_ip": "destination_ip",
        "time": "timestamp",
        "date_time": "timestamp",
        "protocol": "protocol",
        "info": "info",
    }
    return aliases.get(cleaned, cleaned or f"column_{index + 1}")


def _timestamp_from_text(line: str) -> str:
    match = TIME_RE.search(line)
    if match:
        return match.group(1).replace(" ", "T")
    return datetime.now(timezone.utc).isoformat()


def _protocol_from_text(text: str) -> str:
    upper = text.upper()
    for protocol in ("HTTP", "HTTPS", "TCP", "UDP", "DNS", "ICMP", "TLS", "SSH"):
        if protocol in upper:
            return protocol
    return "UNKNOWN"


def _severity_from_text(text: str) -> str:
    upper = text.upper()
    if any(token in upper for token in ("CRITICAL", "MALWARE", "C2", "DDOS", "SYN FLOOD")):
        return "critical"
    if any(token in upper for token in ("FAILED", "SCAN", "SUSPICIOUS", "RST", "504", "REDIRECT")):
        return "high"
    if any(token in upper for token in ("WARN", "DENY", "BLOCK")):
        return "medium"
    return "info"


def _info_from_text(line: str) -> str:
    return line[:500]
